from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import distinct, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app import crud
from app.database import get_db
from app.enums import ActionRequired, PipelineStage, Priority
from app.models import Candidate
from app.schemas import CalendarResponse, CandidateCreate, CandidateRead, CandidateUpdate, MetadataResponse

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["candidates"])

EXPORT_COLUMNS = [
    "id",
    "name",
    "position",
    "pipeline_stage",
    "action_required",
    "priority",
    "next_interview_datetime",
    "expected_joining_date",
    "current_round",
    "notes",
    "created_at",
    "updated_at",
]
ALLOWED_IMPORT_COLUMNS = set(EXPORT_COLUMNS)
UPSERT_COLUMNS = {
    "name",
    "position",
    "pipeline_stage",
    "action_required",
    "priority",
    "next_interview_datetime",
    "expected_joining_date",
    "current_round",
    "notes",
}


def _normalize_header(value: str) -> str:
    return value.strip().lower()


def _clean_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    cleaned = str(value).strip()
    return cleaned if cleaned else None


def _normalize_enum(value: str | None) -> str | None:
    if value is None:
        return None
    return value.strip().upper()


def _parse_upload_rows(upload_file: UploadFile) -> tuple[list[str], list[dict[str, Any]]]:
    filename = (upload_file.filename or "").lower()

    if filename.endswith(".csv"):
        content = upload_file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            raise ValueError("Import file has no header row")
        rows = list(reader)
        return list(reader.fieldnames), rows

    if filename.endswith(".xlsx"):
        workbook = load_workbook(upload_file.file, data_only=True)
        worksheet = workbook.active
        raw_headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if not raw_headers:
            raise ValueError("Import file has no header row")

        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        rows: list[dict[str, Any]] = []
        for row_cells in worksheet.iter_rows(min_row=2, values_only=True):
            row_data: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row_data[header] = row_cells[index] if index < len(row_cells) else None
            rows.append(row_data)
        return headers, rows

    raise ValueError("Only .csv and .xlsx files are supported")


def _extract_candidate_data(
    normalized_row: dict[str, str | None],
    defaults: dict[str, Any],
    is_update: bool,
) -> dict[str, Any]:
    prepared: dict[str, Any] = {}

    for field in UPSERT_COLUMNS:
        if is_update and field not in normalized_row:
            continue

        raw_value = normalized_row.get(field)
        if raw_value is None and not is_update:
            raw_value = defaults.get(field)

        if field in {"pipeline_stage", "action_required", "priority"}:
            raw_value = _normalize_enum(raw_value)

        if field == "notes" and raw_value is None:
            raw_value = ""

        if raw_value is not None:
            prepared[field] = raw_value

    if not is_update:
        prepared.setdefault("name", defaults["name"])
        prepared.setdefault("position", defaults["position"])
        prepared.setdefault("pipeline_stage", defaults["pipeline_stage"])
        prepared.setdefault("action_required", defaults["action_required"])
        prepared.setdefault("priority", defaults["priority"])
        prepared.setdefault("notes", "")

    return prepared


def _row_is_empty(normalized_row: dict[str, str | None]) -> bool:
    for key, value in normalized_row.items():
        if key in {"created_at", "updated_at"}:
            continue
        if value not in {None, ""}:
            return False
    return True


@router.get("/metadata", response_model=MetadataResponse)
def get_metadata(db: Session = Depends(get_db)) -> MetadataResponse:
    positions = list(db.scalars(select(distinct(Candidate.position)).order_by(Candidate.position.asc())).all())
    return MetadataResponse(
        pipeline_stage_values=list(PipelineStage),
        action_required_values=list(ActionRequired),
        priority_values=list(Priority),
        positions=positions,
    )


@router.get("/candidates", response_model=list[CandidateRead])
def get_candidates(
    position: str | None = None,
    pipeline_stage: PipelineStage | None = None,
    action_required: ActionRequired | None = None,
    next_interview_from: datetime | None = None,
    next_interview_to: datetime | None = None,
    expected_joining_from: date | None = None,
    expected_joining_to: date | None = None,
    sort_by: Literal["priority", "next_interview_datetime", "expected_joining_date", "updated_at", "name"] = "updated_at",
    sort_order: Literal["asc", "desc"] = "desc",
    db: Session = Depends(get_db),
) -> list[CandidateRead]:
    candidates = crud.list_candidates(
        db,
        position=position,
        pipeline_stage=pipeline_stage.value if pipeline_stage else None,
        action_required=action_required,
        next_interview_from=next_interview_from,
        next_interview_to=next_interview_to,
        expected_joining_from=expected_joining_from,
        expected_joining_to=expected_joining_to,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return [CandidateRead.model_validate(candidate) for candidate in candidates]


@router.get("/candidates/{candidate_id}", response_model=CandidateRead)
def get_candidate(candidate_id: int, db: Session = Depends(get_db)) -> CandidateRead:
    candidate = crud.get_candidate(db, candidate_id)
    if candidate is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")
    return CandidateRead.model_validate(candidate)


@router.post("/candidates", response_model=CandidateRead, status_code=status.HTTP_201_CREATED)
def create_candidate(payload: CandidateCreate, db: Session = Depends(get_db)) -> CandidateRead:
    try:
        candidate = crud.create_candidate(db, payload)
    except IntegrityError as exc:
        logger.warning("Failed to create candidate due to integrity error: %s", exc)
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid candidate data") from exc
    return CandidateRead.model_validate(candidate)


@router.patch("/candidates/{candidate_id}", response_model=CandidateRead)
def patch_candidate(candidate_id: int, payload: CandidateUpdate, db: Session = Depends(get_db)) -> CandidateRead:
    candidate = crud.get_candidate(db, candidate_id)
    if candidate is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

    try:
        updated = crud.update_candidate(db, candidate, payload)
    except IntegrityError as exc:
        logger.warning("Failed to update candidate %s due to integrity error: %s", candidate_id, exc)
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid update payload") from exc

    return CandidateRead.model_validate(updated)


@router.delete("/candidates/{candidate_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def remove_candidate(candidate_id: int, db: Session = Depends(get_db)) -> Response:
    candidate = crud.get_candidate(db, candidate_id)
    if candidate is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

    crud.delete_candidate(db, candidate)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/candidates-clear")
def clear_candidate_data(db: Session = Depends(get_db)) -> dict[str, int]:
    deleted = crud.clear_candidates(db)
    return {"deleted": deleted}


@router.get("/candidates-export")
def export_candidates(
    file_format: Literal["csv"] = Query(default="csv", alias="format"),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    candidates = crud.list_candidates(db, sort_by="updated_at", sort_order="desc")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()

    for candidate in candidates:
        writer.writerow(
            {
                "id": candidate.id,
                "name": candidate.name,
                "position": candidate.position,
                "pipeline_stage": candidate.pipeline_stage.value,
                "action_required": candidate.action_required.value,
                "priority": candidate.priority.value,
                "next_interview_datetime": candidate.next_interview_datetime.isoformat() if candidate.next_interview_datetime else "",
                "expected_joining_date": candidate.expected_joining_date.isoformat() if candidate.expected_joining_date else "",
                "current_round": candidate.current_round or "",
                "notes": candidate.notes,
                "created_at": candidate.created_at.isoformat() if candidate.created_at else "",
                "updated_at": candidate.updated_at.isoformat() if candidate.updated_at else "",
            }
        )

    content = output.getvalue().encode("utf-8")
    filename = f"candidates_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/candidates-import")
def import_candidates(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict[str, int]:
    defaults: dict[str, Any] = {
        "name": "Unknown Candidate",
        "position": "Unknown Position",
        "pipeline_stage": PipelineStage.TO_BE_SCHEDULED.value,
        "action_required": ActionRequired.SCHEDULE.value,
        "priority": Priority.MEDIUM.value,
    }

    try:
        raw_headers, rows = _parse_upload_rows(file)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    normalized_headers = [_normalize_header(header) for header in raw_headers if header]
    unknown_columns = sorted(set(normalized_headers) - ALLOWED_IMPORT_COLUMNS)
    if unknown_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported columns found: {', '.join(unknown_columns)}",
        )

    created = 0
    updated = 0
    skipped = 0

    try:
        for row_index, row in enumerate(rows, start=2):
            normalized_row = {
                _normalize_header(str(key)): _clean_cell(value)
                for key, value in row.items()
                if key is not None and str(key).strip()
            }

            if _row_is_empty(normalized_row):
                skipped += 1
                continue

            candidate_id: int | None = None
            if "id" in normalized_row and normalized_row["id"] is not None:
                try:
                    candidate_id = int(str(normalized_row["id"]))
                except ValueError as exc:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Row {row_index}: invalid id value",
                    ) from exc

            target = crud.get_candidate(db, candidate_id) if candidate_id is not None else None

            if target is None:
                create_data = _extract_candidate_data(normalized_row, defaults=defaults, is_update=False)
                create_payload = CandidateCreate(
                    name=create_data["name"],
                    position=create_data["position"],
                    priority=create_data["priority"],
                    notes=create_data.get("notes", ""),
                )
                target = Candidate(
                    name=create_payload.name,
                    position=create_payload.position,
                    priority=create_payload.priority,
                    notes=create_payload.notes,
                )
                db.add(target)
                db.flush()

                patch_data = {
                    key: value
                    for key, value in create_data.items()
                    if key in UPSERT_COLUMNS and key not in {"name", "position", "priority", "notes"}
                }
                if patch_data:
                    patch_payload = CandidateUpdate(**patch_data)
                    for field, value in patch_payload.model_dump(exclude_unset=True).items():
                        setattr(target, field, value)

                    crud.apply_stage_defaults(
                        target,
                        stage_was_updated="pipeline_stage" in patch_data,
                        action_was_updated="action_required" in patch_data,
                    )

                created += 1
                continue

            update_data = _extract_candidate_data(normalized_row, defaults=defaults, is_update=True)
            if not update_data:
                skipped += 1
                continue

            update_payload = CandidateUpdate(**update_data)
            update_values = update_payload.model_dump(exclude_unset=True)

            for field, value in update_values.items():
                setattr(target, field, value)

            crud.apply_stage_defaults(
                target,
                stage_was_updated="pipeline_stage" in update_values,
                action_was_updated="action_required" in update_values,
            )
            db.add(target)
            updated += 1

        db.commit()
    except ValidationError as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Import validation error: {exc}") from exc
    except HTTPException:
        db.rollback()
        raise
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Import failed due to invalid data") from exc

    return {"created": created, "updated": updated, "skipped": skipped}


@router.get("/action-items", response_model=list[CandidateRead])
def get_action_items(db: Session = Depends(get_db)) -> list[CandidateRead]:
    candidates = crud.list_action_items(db)
    return [CandidateRead.model_validate(candidate) for candidate in candidates]


@router.get("/calendar", response_model=CalendarResponse)
def get_calendar(
    window: Literal["today", "week"] = Query(default="today"),
    include_joining: bool = True,
    db: Session = Depends(get_db),
) -> CalendarResponse:
    interviews, joinings = crud.get_calendar_candidates(db, window=window, include_joining=include_joining)
    return CalendarResponse(
        interviews=[CandidateRead.model_validate(candidate) for candidate in interviews],
        joinings=[CandidateRead.model_validate(candidate) for candidate in joinings],
    )
